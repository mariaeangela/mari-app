# Lê D:\vida financeira.xlsx e extrai, por mês, os itens de cada categoria
# (descrição + valor), resolvendo as fórmulas J/M feitas à mão.
# Valida que a soma dos itens bate com o total da categoria.
import openpyxl, re, json, sys

XLSX = r'D:\vida financeira.xlsx'
MONTHS = ['jan26','Fev26','Mar26','Abr26','Mai26','Jun26','Jul26']
# rótulos canônicos das 13 categorias (linha J/K 3..15)
CAT_ROWS = list(range(3, 16))  # J3..J15
COL = {'B':2,'C':3,'D':4,'E':5,'G':7,'H':8,'J':10,'K':11,'M':13,'N':14}

wbf = openpyxl.load_workbook(XLSX, data_only=False)
wbv = openpyxl.load_workbook(XLSX, data_only=True)

def cell_f(ws, col, row):
    return ws.cell(row, col).value
def cell_v(ws, wsv, col, row):
    return wsv.cell(row, col).value

CELLREF = re.compile(r"^([A-Z]{1,2})(\d+)$")

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return 0.0

class Resolver:
    def __init__(self, ws, wsv):
        self.ws, self.wsv = ws, wsv
        self.depth = 0
    def col_letter_to_idx(self, l):
        idx = 0
        for ch in l: idx = idx*26 + (ord(ch)-64)
        return idx
    def cell_label_value(self, col_l, row):
        ci = self.col_letter_to_idx(col_l)
        if ci == COL['D']:
            desc = self.ws.cell(row, COL['C']).value
            val = num(self.wsv.cell(row, COL['D']).value)
            card = self.ws.cell(row, COL['B']).value
            return (str(desc).strip() if desc else '(sem descrição)', val, card)
        if ci == COL['M']:
            desc = self.ws.cell(row, COL['N']).value
            # M pode ser fórmula (evento) — pega valor calculado
            val = num(self.wsv.cell(row, COL['M']).value)
            f = self.ws.cell(row, COL['M']).value
            if isinstance(f, str) and f.startswith('='):
                # evento agregado: mantém como 1 item com o rótulo do N
                return (str(desc).strip() if desc else '(evento)', val, None)
            return (str(desc).strip() if desc else '(sem descrição)', val, None)
        if ci == COL['J']:
            # recursa na fórmula da célula J (ex.: streaming J18)
            return None  # tratado no expand
        # outras colunas: usa valor cru
        val = num(self.wsv.cell(row, ci).value)
        return ('(?)', val, None)

    def expand(self, formula, sign=1.0):
        """Retorna lista de (desc, valor, card) folhas de uma fórmula aditiva."""
        if formula is None: return []
        if isinstance(formula, (int, float)):
            return [('(ajuste)', sign*float(formula), None)] if formula else []
        s = str(formula).strip()
        if s.startswith('='): s = s[1:]
        return self._parse_sum(s, sign)

    def _split_top(self, s):
        """Divide por +/- no nível de parênteses 0. Retorna [(sign, term)]."""
        terms, depth, cur, sign = [], 0, '', 1.0
        i = 0
        # sinal inicial
        while i < len(s):
            ch = s[i]
            if ch == '(' : depth += 1; cur += ch
            elif ch == ')': depth -= 1; cur += ch
            elif ch in '+-' and depth == 0:
                if cur.strip(): terms.append((sign, cur.strip()))
                sign = 1.0 if ch == '+' else -1.0
                cur = ''
            else: cur += ch
            i += 1
        if cur.strip(): terms.append((sign, cur.strip()))
        return terms

    def _parse_sum(self, s, outer_sign):
        out = []
        for sgn, term in self._split_top(s):
            out += self._parse_term(term, outer_sign*sgn)
        return out

    def _parse_term(self, term, sign):
        term = term.strip()
        if not term: return []
        # SUM(...)
        m = re.match(r'^SUM\((.*)\)$', term, re.I)
        if m:
            inner = m.group(1)
            items = []
            for part in self._split_commas(inner):
                items += self._parse_range(part.strip(), sign)
            return items
        # parênteses externos
        if term.startswith('(') and term.endswith(')') and self._balanced(term[1:-1]):
            return self._parse_sum(term[1:-1], sign)
        # multiplicação/divisão: cell*factor, cell/factor, factor*cell
        m = re.match(r'^([A-Z]{1,2}\d+)\s*([*/])\s*([\d.]+)$', term)
        if m:
            f = float(m.group(3)); factor = f if m.group(2)=='*' else 1.0/f
            return self._leaf(m.group(1), sign*factor)
        m = re.match(r'^([\d.]+)\s*\*\s*([A-Z]{1,2}\d+)$', term)
        if m:
            return self._leaf(m.group(2), sign*float(m.group(1)))
        # célula simples
        m = CELLREF.match(term)
        if m:
            return self._leaf(term, sign)
        # número puro
        try:
            v = float(term)
            return [('(ajuste)', sign*v, None)] if v else []
        except ValueError:
            # cross-sheet ou algo não previsto: ignora (não ocorre nas categorias)
            return [('(?%s)' % term, 0.0, None)]

    def _leaf(self, ref, sign):
        m = CELLREF.match(ref)
        col_l, row = m.group(1), int(m.group(2))
        ci = self.col_letter_to_idx(col_l)
        if ci == COL['J']:
            f = self.ws.cell(row, COL['J']).value
            return self.expand(f, sign)
        lv = self.cell_label_value(col_l, row)
        if lv is None: return []
        desc, val, card = lv
        return [(desc, sign*val, card)]

    def _parse_range(self, part, sign):
        if ':' in part:
            a, b = part.split(':')
            ma, mb = CELLREF.match(a), CELLREF.match(b)
            col_l = ma.group(1); r1, r2 = int(ma.group(2)), int(mb.group(2))
            out = []
            for r in range(r1, r2+1):
                out += self._leaf(f'{col_l}{r}', sign)
            return out
        return self._leaf(part, sign)

    def _split_commas(self, s):
        parts, depth, cur = [], 0, ''
        for ch in s:
            if ch=='(' : depth+=1; cur+=ch
            elif ch==')': depth-=1; cur+=ch
            elif ch==',' and depth==0: parts.append(cur); cur=''
            else: cur+=ch
        if cur: parts.append(cur)
        return parts

    def _balanced(self, s):
        d=0
        for ch in s:
            if ch=='(':d+=1
            elif ch==')':
                d-=1
                if d<0: return False
        return d==0

result = {}
print(f"{'mês':7} {'categoria':12} {'total':>10} {'itens':>10} {'resid':>8}  n")
for m in MONTHS:
    ws, wsv = wbf[m], wbv[m]
    R = Resolver(ws, wsv)
    cats = []
    for r in CAT_ROWS:
        label = ws.cell(r, COL['K']).value
        if not label: continue
        label = str(label).strip()
        jf = ws.cell(r, COL['J']).value
        jv = num(wsv.cell(r, COL['J']).value)
        items = R.expand(jf)
        # agrupa itens iguais (mesma desc) somando
        agg = {}
        order = []
        for desc, val, card in items:
            if desc not in agg: agg[desc] = 0.0; order.append(desc)
            agg[desc] += val
        leaves = [(d, round(agg[d], 2)) for d in order if abs(agg[d]) > 0.005]
        soma = round(sum(v for _, v in leaves), 2)
        resid = round(jv - soma, 2)
        print(f"{m:7} {label:12} {jv:10.2f} {soma:10.2f} {resid:8.2f}  {len(leaves)}")
        cats.append({'categoria': label, 'total': round(jv,2), 'itens': leaves, 'residual': resid})
    result[m] = cats

with open(r'C:\Users\maria\Downloads\diagonal-claudecode\diagonal\scripts\gastos_import.json','w',encoding='utf-8') as fp:
    json.dump(result, fp, ensure_ascii=False, indent=1)
print('\nOK -> scripts/gastos_import.json')
