# Audita quais transações do Excel NÃO entram em nenhuma categoria (J3:J15).
# Uma transação (coluna D = cartão; coluna M = bloco "gastos mês") está "coberta" se é
# referenciada direto por uma categoria, OU indiretamente via uma célula M que é fórmula.
import openpyxl, re

XLSX = r'D:\vida financeira.xlsx'
MONTHS = ['jan26','Fev26','Mar26','Abr26','Mai26','Jun26','Jul26']
CATROWS = range(3, 16)  # J3..J15
COL = {'B':2,'C':3,'D':4,'E':5,'J':10,'K':11,'M':13,'N':14}

wbf = openpyxl.load_workbook(XLSX, data_only=False)
wbv = openpyxl.load_workbook(XLSX, data_only=True)

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return 0.0

def extract_refs(f):
    refs = set()
    if not isinstance(f, str) or not f.startswith('='):
        return refs
    s = f[1:]
    for m in re.finditer(r'([A-Z]{1,2})(\d+):([A-Z]{1,2})(\d+)', s):
        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        if c1 == c2:
            for r in range(r1, r2 + 1): refs.add((c1, r))
    s2 = re.sub(r'[A-Z]{1,2}\d+:[A-Z]{1,2}\d+', '', s)
    for m in re.finditer(r'([A-Z]{1,2})(\d+)', s2):
        refs.add((m.group(1), int(m.group(2))))
    return refs

grand_leak = 0.0
grand_reemb = 0.0
for mo in MONTHS:
    ws, wv = wbf[mo], wbv[mo]
    # coleta refs das categorias, recursando em células J
    work = set()
    for r in CATROWS:
        work |= extract_refs(ws.cell(r, COL['J']).value)
    D_direct, M_ref, seenJ = set(), set(), set()
    stack = list(work)
    while stack:
        c, row = stack.pop()
        if c == 'D': D_direct.add(row)
        elif c == 'M': M_ref.add(row)
        elif c == 'J':
            if row not in seenJ:
                seenJ.add(row)
                stack.extend(extract_refs(ws.cell(row, COL['J']).value))
        # outras colunas: ignora
    D_via_M = set()
    for mrow in M_ref:
        for c, row in extract_refs(ws.cell(mrow, COL['M']).value):
            if c == 'D': D_via_M.add(row)
    covered_D = D_direct | D_via_M

    # varre transações de cartão (coluna D) não cobertas
    leaks = []
    reembs = []
    maxr = 0
    for r in range(6, 1002):
        if ws.cell(r, COL['D']).value is not None: maxr = r
    for r in range(6, maxr + 1):
        raw = ws.cell(r, COL['D']).value
        if raw is None: continue
        val = num(wv.cell(r, COL['D']).value)
        if val == 0: continue
        if r in covered_D: continue
        desc = ws.cell(r, COL['C']).value or '(sem desc)'
        status = (ws.cell(r, COL['E']).value or '').strip()
        if status.upper() == 'REEMBOLSO':
            reembs.append((r, str(desc).strip(), val, status)); continue
        leaks.append((r, str(desc).strip(), val, status))
    # varre bloco M não referenciado
    for r in range(3, 40):
        if ws.cell(r, COL['M']).value is None: continue
        val = num(wv.cell(r, COL['M']).value)
        if val == 0: continue
        if r in M_ref: continue
        desc = ws.cell(r, COL['N']).value or '(sem desc)'
        leaks.append(('M%d' % r, str(desc).strip(), val, 'bloco M'))

    parcelas = [x for x in leaks if 'parcela' in (x[3] or '').lower()]
    outros = [x for x in leaks if 'parcela' not in (x[3] or '').lower()]
    tot_parc = sum(v for _, _, v, _ in parcelas)
    tot_outros = sum(v for _, _, v, _ in outros)
    tot_reemb = sum(v for _, _, v, _ in reembs)
    grand_leak += tot_parc; grand_reemb += tot_reemb
    globals()['G_OUTROS'] = globals().get('G_OUTROS', 0.0) + tot_outros
    print(f'\n===== {mo}  |  parcelas fora de cat: R$ {tot_parc:,.2f}  |  OUTROS (possível esquecimento): R$ {tot_outros:,.2f}  |  reembolsos: R$ {tot_reemb:,.2f}')
    if outros:
        print('   -- OUTROS (nao-parcela, nao-reembolso):')
        for cell, desc, val, st in sorted(outros, key=lambda x: -x[2]):
            print(f'      {desc:30s} R$ {val:9.2f}   [{cell}]{" · "+st if st else ""}')

print(f'\n########  parcelas fora de categoria (jan-jul): R$ {grand_leak:,.2f}  (corretas: a compra cheia ja foi categorizada)')
print(f'########  OUTROS possiveis esquecimentos: R$ {globals().get("G_OUTROS",0):,.2f}')
print(f'########  reembolsos (excluidos de proposito): R$ {grand_reemb:,.2f}')
